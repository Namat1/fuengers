import streamlit as st
import pandas as pd
import io
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

st.title("Zulage Füngers")

uploaded_files = st.file_uploader("Excel-Dateien hochladen", type=["xlsx"], accept_multiple_files=True)

german_months = {
    1: "Januar", 2: "Februar", 3: "März", 4: "April",
    5: "Mai", 6: "Juni", 7: "Juli", 8: "August",
    9: "September", 10: "Oktober", 11: "November", 12: "Dezember"
}

if uploaded_files:
    eintraege = []

    for file in uploaded_files:
        try:
            df = pd.read_excel(file, sheet_name="Touren", header=None)
            df = df.iloc[4:]
            df.columns = range(df.shape[1])

            for _, row in df.iterrows():
                kommentar = str(row[15]) if 15 in row and pd.notnull(row[15]) else ""
                name = row[3] if 3 in row else None
                vorname = row[4] if 4 in row else None
                datum = pd.to_datetime(row[14], errors='coerce') if 14 in row else None

                if (
                    "füngers" in kommentar.lower()
                    and pd.notnull(name)
                    and pd.notnull(vorname)
                    and pd.notnull(datum)
                ):
                    kw = datum.isocalendar().week
                    datum_kw = datum.strftime("%d.%m.%Y") + f" (KW {kw})"
                    monat_index = datum.month
                    jahr = datum.year
                    monat_name = german_months[monat_index]
                    eintraege.append({
                        "Nachname": name,
                        "Vorname": vorname,
                        "DatumKW": datum_kw,
                        "Kommentar": kommentar,
                        "Verdienst": 20,
                        "Monat": f"{monat_index:02d}-{jahr}_{monat_name} {jahr}"
                    })

        except Exception as e:
            st.error(f"Fehler in Datei {file.name}: {e}")

    if eintraege:
        df_gesamt = pd.DataFrame(eintraege)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            for monat_key in sorted(df_gesamt["Monat"].unique()):
                df_monat = df_gesamt[df_gesamt["Monat"] == monat_key]
                zeilen = []
                for (nach, vor), gruppe in df_monat.groupby(["Nachname", "Vorname"]):
                    zeilen.append([f"{vor} {nach}", "", ""])
                    zeilen.append(["Datum", "Kommentar", "Verdienst"])
                    for _, r in gruppe.iterrows():
                        zeilen.append([r["DatumKW"], r["Kommentar"], r["Verdienst"]])
                    zeilen.append(["Gesamt", "", gruppe["Verdienst"].sum()])
                    zeilen.append(["", "", ""])

                monatsgesamt = df_monat["Verdienst"].sum()

                df_sheet = pd.DataFrame(zeilen, columns=["Spalte A", "Spalte B", "Spalte C"])
                sheet_name = monat_key.split("_")[1][:31]
                df_sheet.to_excel(writer, index=False, sheet_name=sheet_name)

                sheet = writer.sheets[sheet_name]
                sheet.row_dimensions[1].hidden = True

                # Moderne Farbpalette
                thin = Border(
                    left=Side(style='thin', color='CCCCCC'),
                    right=Side(style='thin', color='CCCCCC'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='thin', color='CCCCCC')
                )
                
                medium_border = Border(
                    left=Side(style='medium', color='1F4E78'),
                    right=Side(style='medium', color='1F4E78'),
                    top=Side(style='medium', color='1F4E78'),
                    bottom=Side(style='medium', color='1F4E78')
                )

                name_fill = PatternFill("solid", fgColor="4472C4")  # Mittelblau
                header_fill = PatternFill("solid", fgColor="D9E2F3")  # Hellblau
                total_fill = PatternFill("solid", fgColor="70AD47")  # Grün
                data_fill_white = PatternFill("solid", fgColor="FFFFFF")  # Weiß
                data_fill_light = PatternFill("solid", fgColor="F8F9FA")  # Hellgrau

                monatsgesamt_row = None
                alternate_row = False

                for row in sheet.iter_rows():
                    row_idx = row[0].row
                    val = str(row[0].value).strip().lower() if row[0].value else ""
                    is_name_row = (
                        str(row[0].value).strip() != ""
                        and (row[1].value is None or row[1].value == "")
                        and (row[2].value is None or row[2].value == "")
                    )

                    # Namenszeile (z.B. "Max Mustermann")
                    if is_name_row:
                        alternate_row = False
                        for cell in row:
                            cell.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            cell.fill = name_fill
                            cell.border = medium_border
                        sheet.row_dimensions[row_idx].height = 22

                    # Header-Zeile (Datum, Kommentar, Verdienst)
                    elif row[0].value == "Datum":
                        alternate_row = False
                        for cell in row:
                            cell.font = Font(name="Calibri", bold=True, size=10, color="1F4E78")
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            cell.fill = header_fill
                            cell.border = thin
                        sheet.row_dimensions[row_idx].height = 20

                    # Gesamt-Zeile pro Person
                    elif val == "gesamt":
                        alternate_row = False
                        for cell in row:
                            cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                            cell.fill = total_fill
                            cell.border = medium_border
                        sheet.row_dimensions[row_idx].height = 20

                    # Leere Trennzeile
                    elif val == "":
                        for cell in row:
                            cell.font = Font(name="Calibri", size=11)
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                            cell.border = None
                        if monatsgesamt_row is None:
                            monatsgesamt_row = row_idx + 1
                        alternate_row = False

                    # Datenzeilen (alternierende Farben)
                    else:
                        fill_color = data_fill_white if alternate_row else data_fill_light
                        for cell in row:
                            cell.font = Font(name="Calibri", size=10, color="2C3E50")
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                            cell.fill = fill_color
                            cell.border = thin
                        sheet.row_dimensions[row_idx].height = 20
                        alternate_row = not alternate_row

                    # Format Spalte C (Verdienst) mit Euro
                    verdienst_cell = row[2]
                    try:
                        if isinstance(verdienst_cell.value, (float, int)):
                            verdienst_cell.number_format = '#,##0.00 €'
                            verdienst_cell.alignment = Alignment(horizontal="right", vertical="center")
                            # Grüne Schrift für positive Beträge in Datenzeilen
                            if verdienst_cell.value > 0 and not (is_name_row or val == "gesamt"):
                                verdienst_cell.font = Font(name="Calibri", size=10, color="70AD47", bold=True)
                    except:
                        pass

                # Monatsgesamt in Spalte E/F schreiben
                if monatsgesamt_row:
                    cell_text = sheet.cell(row=monatsgesamt_row, column=5)  # E
                    cell_text.value = "Monatsgesamt:"
                    cell_text.font = Font(name="Calibri", bold=True, size=13, color="1F4E78")
                    cell_text.alignment = Alignment(horizontal="right", vertical="center")
                    cell_text.fill = PatternFill("solid", fgColor="D9E2F3")
                    cell_text.border = medium_border

                    cell_sum = sheet.cell(row=monatsgesamt_row, column=6)  # F
                    cell_sum.value = monatsgesamt
                    cell_sum.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
                    cell_sum.number_format = '#,##0.00 €'
                    cell_sum.alignment = Alignment(horizontal="right", vertical="center")
                    cell_sum.fill = PatternFill("solid", fgColor="70AD47")
                    cell_sum.border = medium_border
                    
                    sheet.row_dimensions[monatsgesamt_row].height = 24

                # Spaltenbreiten mit Mindestbreiten
                column_min_widths = {
                    1: 25,  # Spalte A (Datum/Name)
                    2: 35,  # Spalte B (Kommentar - braucht mehr Platz)
                    3: 18,  # Spalte C (Verdienst)
                    5: 20,  # Spalte E (Monatsgesamt-Label)
                    6: 20   # Spalte F (Monatsgesamt-Betrag)
                }

                for col_idx, col_cells in enumerate(sheet.columns, start=1):
                    max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col_cells)
                    col_letter = get_column_letter(col_cells[0].column)
                    
                    # Berechne Breite mit großzügigerem Puffer
                    calculated_width = int(max_len * 1.2) + 4
                    
                    # Verwende Mindestbreite falls definiert
                    min_width = column_min_widths.get(col_idx, 12)
                    adjusted_width = max(calculated_width, min_width)
                    
                    # Maximalbreite begrenzen
                    adjusted_width = min(adjusted_width, 70)
                    
                    sheet.column_dimensions[col_letter].width = adjusted_width

                # Freeze Panes für bessere Navigation
                sheet.freeze_panes = "A3"

        st.download_button("Excel-Datei herunterladen", output.getvalue(), file_name="Füngers_Monatsauswertung.xlsx")

    else:
        st.warning("Keine gültigen Füngers-Zulagen gefunden.")
